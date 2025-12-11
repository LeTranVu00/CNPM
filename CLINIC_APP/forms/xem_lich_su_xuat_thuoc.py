from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QMessageBox, QDateEdit, QDialog, QSizePolicy
)
from PyQt5.QtCore import Qt, QDate
from database import get_xuat_thuoc_history, get_connection, get_prescriptions_for_patient


class XemLichSuXuatThuoc(QWidget):
    """Form để xem lịch sử xuất thuốc - dành cho admin."""
    def __init__(self, parent=None, username=None):
        super().__init__(parent)
        self.username = username
        self.setWindowTitle("Lịch sử Xuất thuốc")
        self.init_ui()
        self.load_history()

    def init_ui(self):
        layout = QVBoxLayout(self)
        header = QLabel("LỊCH SỬ XUẤT THUỐC")
        header.setStyleSheet('font-weight:bold; font-size:14pt; color:#1565c0')
        layout.addWidget(header)

        # Các điều khiển lọc
        controls = QHBoxLayout()
        lbl_from = QLabel("Từ ngày:")
        self.date_from = QDateEdit()
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_from.setCalendarPopup(True)
        
        lbl_to = QLabel("Đến ngày:")
        self.date_to = QDateEdit()
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        
        btn_filter = QPushButton("Lọc")
        btn_filter.clicked.connect(self.load_history)
        
        btn_refresh = QPushButton("Tải lại")
        btn_refresh.clicked.connect(self.load_history)
        
        controls.addWidget(lbl_from)
        controls.addWidget(self.date_from)
        controls.addWidget(lbl_to)
        controls.addWidget(self.date_to)
        controls.addWidget(btn_filter)
        controls.addWidget(btn_refresh)
        controls.addStretch()
        layout.addLayout(controls)

        # Bảng lịch sử
        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels([
            "Đơn ID", "Bác sĩ kê", "Bệnh nhân", "Số CCCD", 
            "Xuất bởi", "Thời gian xuất", "Ghi chú", "Xem"
        ])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setMinimumHeight(400)
        self.table.resizeColumnsToContents()
        layout.addWidget(self.table, 1)  # Đặt hệ số giãn cho bảng là 1 để lấp đầy không gian

        # Các nút hành động
        actions = QHBoxLayout()
        self.btn_export = QPushButton("Xuất Excel")
        self.btn_export.clicked.connect(self.export_to_excel)
        actions.addStretch()
        actions.addWidget(self.btn_export)
        layout.addLayout(actions)

    def load_history(self):
        """Load history from database."""
        try:
            history = get_xuat_thuoc_history(limit=500)
            self.table.setRowCount(0)
            
            date_from = self.date_from.date().toString('yyyy-MM-dd')
            date_to = self.date_to.date().toString('yyyy-MM-dd')
            
            for row_data in history:
                # Lọc theo khoảng ngày
                thoi_gian = row_data[6]  # thoi_gian_xuat
                if thoi_gian:
                    date_part = thoi_gian.split(' ')[0]  # Extract date from datetime
                    if not (date_from <= date_part <= date_to):
                        continue
                
                r = self.table.rowCount()
                self.table.insertRow(r)
                
                # row_data: id, don_thuoc_id, dac_si, ho_ten_benh_nhan, so_cccd, xuat_boi, thoi_gian_xuat, ghi_chu (dữ liệu hàng)
                # Bỏ qua ID (row_data[0]) và bắt đầu từ don_thuoc_id
                self.table.setItem(r, 0, QTableWidgetItem(str(row_data[1])))  # đơn ID
                self.table.setItem(r, 1, QTableWidgetItem(str(row_data[2] or '')))  # bác sĩ
                self.table.setItem(r, 2, QTableWidgetItem(str(row_data[3] or '')))  # bệnh nhân
                self.table.setItem(r, 3, QTableWidgetItem(str(row_data[4] or '')))  # CCCD
                self.table.setItem(r, 4, QTableWidgetItem(str(row_data[5] or '')))  # xuất bởi
                self.table.setItem(r, 5, QTableWidgetItem(str(row_data[6] or '')))  # thời gian
                self.table.setItem(r, 6, QTableWidgetItem(str(row_data[7] or '')))  # ghi chú
                
                # Thêm nút "Xem chi tiết" ở cột 7
                try:
                    btn = QPushButton("Xem")
                    btn.setToolTip("Xem chi tiết đơn xuất thuốc")
                    # Lưu don_thuoc_id (row_data[1]) để dùng trong callback
                    btn.clicked.connect(lambda _, don_id=row_data[1]: self.open_detail(don_id))
                    self.table.setCellWidget(r, 7, btn)
                except Exception:
                    pass
            
            self.table.resizeColumnsToContents()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi tải lịch sử: {e}")

    def open_detail(self, don_thuoc_id):
        """Hiển thị chi tiết của một đơn xuất thuốc."""
        try:
            conn = get_connection()
            cur = conn.cursor()
            
            # Lấy thông tin đơn thuốc
            cur.execute("""
                SELECT dt.id, dt.phieu_kham_id, dt.ngay_ke, dt.bac_si, dt.so_ngay, dt.tong_tien,
                       pk.benh_nhan_id
                FROM don_thuoc dt
                LEFT JOIN phieu_kham pk ON dt.phieu_kham_id = pk.id
                WHERE dt.id = ?
            """, (don_thuoc_id,))
            don_info = cur.fetchone()
            
            if not don_info:
                QMessageBox.warning(self, "Lỗi", "Không tìm thấy thông tin đơn thuốc.")
                conn.close()
                return
            
            don_id, phieu_kham_id, ngay_ke, bac_si, so_ngay, tong_tien, benh_nhan_id = don_info
            
            # Lấy chi tiết thuốc trong đơn
            cur.execute("""
                SELECT ma_thuoc, ten_thuoc, so_luong, don_vi, lieu_dung, ghi_chu
                FROM chi_tiet_don_thuoc
                WHERE don_thuoc_id = ?
                ORDER BY id
            """, (don_thuoc_id,))
            drug_details = cur.fetchall()
            
            conn.close()
            
            # Tạo dialog hiển thị chi tiết
            dlg = QDialog(self)
            dlg.setWindowTitle(f"Chi tiết Đơn Xuất Thuốc #{don_thuoc_id}")
            dlg.setMinimumSize(800, 500)
            layout = QVBoxLayout(dlg)
            
            # Thông tin chung
            info_text = f"<b>Đơn ID:</b> {don_id} | <b>Ngày kê:</b> {ngay_ke} | <b>Bác sĩ:</b> {bac_si}"
            info_label = QLabel(info_text)
            layout.addWidget(info_label)
            
            # Bảng chi tiết thuốc
            table = QTableWidget(0, 6)
            table.setHorizontalHeaderLabels(["Mã thuốc", "Tên thuốc", "Số lượng", "Đơn vị", "Liều dùng", "Ghi chú"])
            table.setSelectionBehavior(QTableWidget.SelectRows)
            table.setEditTriggers(QTableWidget.NoEditTriggers)
            
            for idx, drug in enumerate(drug_details):
                table.insertRow(idx)
                ma, ten, so_luong, don_vi, lieu_dung, ghi_chu = drug
                table.setItem(idx, 0, QTableWidgetItem(str(ma or "")))
                table.setItem(idx, 1, QTableWidgetItem(str(ten or "")))
                table.setItem(idx, 2, QTableWidgetItem(str(so_luong or "")))
                table.setItem(idx, 3, QTableWidgetItem(str(don_vi or "")))
                table.setItem(idx, 4, QTableWidgetItem(str(lieu_dung or "")))
                table.setItem(idx, 5, QTableWidgetItem(str(ghi_chu or "")))
            
            # Điều chỉnh độ rộng cột
            table.horizontalHeader().setSectionResizeMode(0, table.horizontalHeader().Fixed)
            table.horizontalHeader().setSectionResizeMode(1, table.horizontalHeader().Stretch)
            table.setColumnWidth(0, 100)
            layout.addWidget(table)
            
            # Nút đóng
            btn_close = QPushButton("Đóng")
            btn_close.clicked.connect(dlg.accept)
            layout.addWidget(btn_close)
            
            dlg.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi hiển thị chi tiết: {e}")

    def export_to_excel(self):
        """Export history to Excel file."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Tạo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Lịch sử xuất thuốc"
            
            # Tiêu đề cột
            headers = ["Đơn ID", "Bác sĩ kê", "Bệnh nhân", "Số CCCD", 
                      "Xuất bởi", "Thời gian xuất", "Ghi chú"]
            ws.append(headers)
            
            # Định dạng tiêu đề
            header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Dữ liệu
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
            
            # Điều chỉnh độ rộng cột
            for col in range(1, self.table.columnCount() + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            # Lưu
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"lich_su_xuat_thuoc_{timestamp}.xlsx"
            wb.save(filename)
            QMessageBox.information(self, "Thành công", f"Đã xuất file: {filename}")
        except ImportError:
            QMessageBox.warning(self, "Cảnh báo", "Cần cài đặt openpyxl để xuất Excel.\nChạy: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi xuất Excel: {e}")
